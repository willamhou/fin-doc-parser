"""Excel file parser — converts XLS/XLSX to markdown text."""

from __future__ import annotations

from pathlib import Path


def parse_excel(file_path: Path) -> str:
    """Parse an Excel file and return content as markdown text.

    Tries openpyxl first (xlsx), falls back to xlrd (xls).

    Returns:
        Markdown-formatted table content.
    """
    ext = file_path.suffix.lower()

    if ext == ".csv":
        return _parse_csv(file_path)
    if ext == ".xls":
        return _parse_xls(file_path)
    return _parse_xlsx(file_path)


def _parse_xlsx(file_path: Path) -> str:
    """Parse .xlsx file using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError as err:
        raise ImportError(
            "openpyxl is required for .xlsx files. "
            "Install with: pip install 'fin-doc-parser[excel]'"
        ) from err

    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"## Sheet: {sheet_name}\n")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            parts.append("(empty)\n")
            continue

        # Build markdown table
        header = rows[0]
        col_names = [str(c) if c is not None else "" for c in header]
        parts.append("| " + " | ".join(col_names) + " |")
        parts.append("| " + " | ".join(["---"] * len(col_names)) + " |")

        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("| " + " | ".join(cells) + " |")

        parts.append("")

    wb.close()
    return "\n".join(parts)


def _parse_xls(file_path: Path) -> str:
    """Parse .xls file using xlrd."""
    try:
        import xlrd
    except ImportError as err:
        raise ImportError(
            "xlrd is required for .xls files. Install with: pip install 'fin-doc-parser[excel]'"
        ) from err

    wb = xlrd.open_workbook(str(file_path))
    parts: list[str] = []

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        parts.append(f"## Sheet: {sheet_name}\n")

        if ws.nrows == 0:
            parts.append("(empty)\n")
            continue

        header = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join(["---"] * len(header)) + " |")

        for r in range(1, ws.nrows):
            cells = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
            parts.append("| " + " | ".join(cells) + " |")

        parts.append("")

    return "\n".join(parts)


_CSV_ENCODINGS = ("utf-8-sig", "gbk", "gb2312", "latin-1")


def _parse_csv(file_path: Path) -> str:
    """Parse .csv file with multi-encoding fallback."""
    import csv

    for encoding in _CSV_ENCODINGS:
        try:
            with open(file_path, newline="", encoding=encoding) as f:
                reader = csv.reader(f)
                rows = list(reader)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(
            f"Cannot decode CSV file {file_path.name}. Tried encodings: {', '.join(_CSV_ENCODINGS)}"
        )

    if not rows:
        return "(empty)"

    parts: list[str] = ["## CSV\n"]
    header = rows[0]
    parts.append("| " + " | ".join(header) + " |")
    parts.append("| " + " | ".join(["---"] * len(header)) + " |")

    for row in rows[1:]:
        parts.append("| " + " | ".join(row) + " |")

    return "\n".join(parts)
